# -*- coding: utf-8 -*-
"""
Created on Fri Jul  9 06:57:53 2021

@author: Ilham Miftha Faiz
"""


import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from django.conf import settings
#Preprocessing

def nodin_to_atoll(Nodin,RNC):
    objnodin = Nodin
    objrnc = RNC
    excel_name = objnodin.Nodin.name
    
    filename = objnodin.Nodin.path
    filernc = objrnc.RNC.path
    file_output = "TPASS" + excel_name[14:]
    directory = os.path.join(settings.MEDIA_ROOT, 'Output')
    if not os.path.exists(directory):
        os.makedirs(directory)
    writer = pd.ExcelWriter(os.path.join(directory, file_output), engine='xlsxwriter')
    
    df_rnc = pd.read_excel(filernc, sheet_name="RNC", header=1, engine= 'openpyxl',)
    df_locno = pd.read_excel(filernc, sheet_name="loc no", header=0, engine= 'openpyxl',)
    df_mgw = df_rnc[['RNC Name','MGW']].copy()
    df_rncspc = df_rnc[['RNC Name','RNC SPC']].copy()
    df_tz = df_locno[["Kabupaten","Time Zone"]].copy()
    df_tz['Kabupaten'] = df_tz['Kabupaten'].str.title()
    df_locno1 = df_locno[['Kabupaten','LOC No']].copy()
    df_locno1['Kabupaten'] = df_locno1['Kabupaten'].str.title()
    wb1 = load_workbook(filename)
    
    if "4G" in wb1.sheetnames:
        df_nodin4G = pd.read_excel(filename, sheet_name="4G",header = 1, engine= 'openpyxl',)
        
        
        #count
        df_4Gcount= pd.DataFrame()
        df_4Gcount['NE ID'] = df_nodin4G['NE ID'].copy()
        df_4Gcount = df_4Gcount.groupby(['NE ID'])[['NE ID']].count()
        
        #first for loop processing
        df_raw4 = pd.DataFrame()
        
        for x in range(9):
            ty = x + 1
            df_selection4 = pd.DataFrame()
            df_selection4['NE ID'] = df_4Gcount[df_4Gcount['NE ID'] == ty].index
            if not df_selection4.empty:
                df_temp4G = pd.merge(df_selection4,df_nodin4G, how = 'left' , on = 'NE ID')
                df_raw14 = df_temp4G[["Site_ID", "NE ID" , "SITE NAME\n(CONNECTED)","Cell ID", "CELL NAME\n(CONNECTED)", "Longtitude",
                                    "Latitude", "Province","Kabupaten", "Kecamatan", "Kelurahan", "Address", 'LOCNO', "Branch", "Sales_Cluster", 
                                    "Antenna 1", "Antenna_Height (m)", "AFTER", "MTILT", "ETILT", "TP", "TAL", "TAC", 
                                    "PCI", "E_Node_B_ID", "BAND", "DEARFCN", "UEARFCN","MCC", "MNC"
                                    ]].copy()
                df_raw14.columns = ['site_id', 'ne_id', 'BTS/NodeB/eNodeB_Name','Cell ID','Cell Name', 'Longitude', 'Latitude', 
                                   'Province', 'Kabupaten', 'Kecamatan', 'Kelurahan', 'Address', 'LOC No', 'Branch', 'Cluster SCS','Antenna Type',
                                   'Antenna Height', 'Azimut', 'M-Tilt Antena', 'E-Tilt Antena', 'Tower Provider',"TAL", "TAC", "PCI",
                                   "Enodeb Id","Frequency Band","DL EARFCN", "UL EARFCN","MCC", "MNC"]
                
                df_raw14.insert(0, 'network_type_band', "4G", allow_duplicates=False)
                df_raw14.insert(0, 'vendor', "ZTE", allow_duplicates=False)
                df_raw14.insert(0, 'order_id', np.nan, allow_duplicates=False)
                
                df_raw14.insert(6, 'Site Name OSS', np.nan, allow_duplicates=False)
        
                df_raw14.insert(9, 'SoW', np.nan, allow_duplicates=False)        
        
                df_raw14.insert(18, 'Network Area', "Area 04", allow_duplicates=False)
                df_raw14.insert(18, 'Regional', "R11 Papua Maluku", allow_duplicates=False)
               
                df_raw14.insert(22, 'Config', np.nan, allow_duplicates=False)
                df_raw14.insert(22, 'Type BTS', np.nan, allow_duplicates=False)
                df_raw14.insert(22, 'NE Type', np.nan, allow_duplicates=False)
                df_raw14.insert(22, 'Time Zone', np.nan, allow_duplicates=False)
                
                df_raw14.insert(35, 'Cell Radius (m)', 300, allow_duplicates=False)
                df_raw14.insert(35, 'MOD6', np.nan, allow_duplicates=False)        
                df_raw14.insert(35, 'PRACH',np.nan, allow_duplicates=False)        
                df_raw14.insert(35, 'MOD3', np.nan, allow_duplicates=False)
                
                df_raw14.insert(40, "Enodeb Id (Hexa)", np.nan, allow_duplicates=False)
                
                df_raw14.insert(44, 'P-GW', "TBD", allow_duplicates=False)
                df_raw14.insert(44, 'S-GW', "TBD", allow_duplicates=False)
                df_raw14.insert(44, 'MME', "TBD", allow_duplicates=False)
                
                df_raw14.loc[:,'Remark'] = ty
                
                df_raw4 = pd.concat([df_raw4,df_raw14],axis = 0,ignore_index= True)
        
        #second for loops
        df_raw4['Site Name OSS'] = df_raw4['BTS/NodeB/eNodeB_Name'].str[2:]
        df_raw4['NE Type'] = df_raw4['ne_id'].str[6]
        df_raw4['MOD3'] = df_raw4['PCI'].mod(3)
        df_raw4['MOD6'] = df_raw4['PCI'].mod(6)
        df_raw4['Enodeb Id (Hexa)'] = df_raw4['Enodeb Id'].apply( hex )
        df_raw4['Frequency Band'] = df_raw4['Frequency Band'].str[1:]
        df_raw4['PCI'] = df_raw4['PCI'].fillna(0)
        df_raw4['MOD3'] = df_raw4['MOD3'].fillna(0)
        df_raw4['MOD6'] = df_raw4['MOD6'].fillna(0)
        df_raw4['PCI'] = df_raw4['PCI'].astype(int)
        df_raw4['MOD3'] = df_raw4['MOD3'].astype(int)
        df_raw4['MOD6'] = df_raw4['MOD6'].astype(int)
        
        
        df_4GTPASS = pd.DataFrame()
        
        
        for x in range(9):
            ty = x + 1
            df_selection4 = pd.DataFrame()
            df_selection4 = df_raw4[df_raw4['Remark'] == ty].copy()
            df_colon14 = df_selection4[['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena','PCI','MOD3','MOD6']].copy()
            df_colon4 = pd.DataFrame()
            df_4Gprep = pd.DataFrame()
            if not df_selection4.empty:
                if ty == 1:
                    df_4GTPASS1 = df_selection4.copy()
                    df_4GTPASS = pd.concat([df_4GTPASS,df_4GTPASS1],axis = 0,ignore_index= True)
                else:
                    for z in range(0,len(df_selection4),ty):
                        df_tempc = pd.DataFrame()
                        for y in range(ty):
                            t = z + y
                            data = df_colon14.loc[df_colon14.index[t],['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena','PCI','MOD3','MOD6']].values
                            df_tempc.insert(y, y, data, allow_duplicates=False)
                        if ty == 1:
                            df_tempc['Concat'] = df_tempc[0].astype(str)
                        else:
                            df_tempc['Sementara Concat'] = df_tempc[0].astype(str)
                            for w in range(1,ty):
                                df_tempc['Concat'] = df_tempc['Sementara Concat'] +';'+ df_tempc[w].astype(str)
                                df_tempc['Sementara Concat'] = df_tempc['Concat']
                        df_tempt = df_tempc.transpose()
                        df_tempt.columns =['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena','PCI','MOD3','MOD6']
                        df_tempt.loc['Concat','ne_id'] = df_tempt.loc[0,'ne_id']
                        df_tempc =  df_tempt.loc['Concat',:].copy()
                        df_colon4 = df_colon4.append(df_tempc,ignore_index=True)
                        df_4Gprept = pd.DataFrame(df_selection4.loc[df_selection4.index[(z)],df_selection4.columns].values.reshape(1,-1), columns = df_selection4.columns)
                        df_4Gprep = df_4Gprep.append(df_4Gprept,ignore_index=True)
                        df_colon4 = df_colon4.drop(columns=['ne_id'],axis = 1)
                df_4Gprep.update(df_colon4)
                df_4GTPASS = pd.concat([df_4GTPASS,df_4Gprep],axis = 0,ignore_index= True)
        df_4GTPASS['Config'] = df_4GTPASS['Remark'].copy()
        df_4GTPASS['Config'] = df_4GTPASS['Config'].replace({1:'S1', 2:'S11', 3:'S111', 4:'S1111', 5:'S11111', 6:'S111111'})
        df_4GTPASS['Tower Provider'] = df_4GTPASS['Tower Provider'].replace({"-":"TBD"})
        df_4GTPASS['PRACH'] = df_4GTPASS['ne_id'].str[7]
        df_4GTPASS['PRACH'] = df_4GTPASS['PRACH'].replace({"L":"TA JAUH","T":"TA JAUH","E":"TA DEKAT","R":"TA DEKAT","F":"TA DEKAT"})
        df_4GTPASS['NE Type'] = df_4GTPASS['NE Type'].replace({"M":"Macro","I":"Indoor","H":"Hotel"})
        df_4GTPASS = df_4GTPASS.drop('Remark',axis =1)
        
        df_4GTPASS['Kabupaten'] = df_4GTPASS['Kabupaten'].str.title()
        df_4GTPASS = pd.merge(df_4GTPASS,df_tz, how = 'left' , on = 'Kabupaten')
        df_4GTPASS["Time Zone_x"] = df_4GTPASS["Time Zone_y"]
        df_4GTPASS = df_4GTPASS.rename(columns = {'Time Zone_x': 'Time Zone'}, inplace = False)
        df_4GTPASS = df_4GTPASS.drop("Time Zone_y",axis = 1)
        
        
        sheetname1 = '4G'
        df_4GTPASS.to_excel(writer, sheet_name = sheetname1, index=False)     
    
    
    
    #First Table
    
    #count
    
    if "3G" in wb1.sheetnames:
        df_nodin3G = pd.read_excel(filename, sheet_name="3G",header = 1, engine= 'openpyxl',)
            
        df_3Gcount= pd.DataFrame()
        df_3Gcount['NE ID'] = df_nodin3G['NE ID'].copy()
        df_3Gcount = df_3Gcount.groupby(['NE ID'])[['NE ID']].count()
    
    
        df_raw3 = pd.DataFrame()
    
    
        for x in range(9):
            ty = x + 1    
            df_selection3 = pd.DataFrame()
            df_selection3['NE ID'] = df_3Gcount[df_3Gcount['NE ID'] == ty].index
            
            if not df_selection3.empty:
                df_temp3G = pd.merge(df_selection3,df_nodin3G, how = 'left' , on = 'NE ID')
                df_raw13 = df_temp3G[["SITE ID", "NE ID", "SITE NAME\n(CONNECTED)", "CI", "CELL NAME\n(CONNECTED)",
                                        "Longitude", "Latitude", "PROVINCE", "KOTA_KAB", "KOTA_KEC", "KOTA_KEL", "Address",
                                        "BRANCH", "Sales Cluster", "Antenna Type", "Antenna Height(m)", "Azimuth(°)", "MTILT", 
                                        "ETILT", "TP", "LAC", "RNC Name", "RAC", "URA ID Planning", "SACPC", "Primary\nScrambling\nCode",
                                        "MCC", "MNC", "Carrier"
                                    ]].copy()
                df_raw13.columns = ['site_id', 'ne_id', 'BTS/NodeB/eNodeB_Name','Cell ID', 'Cell Name',
                                    'Longitude', 'Latitude', 'Province', 'Kabupaten', 'Kecamatan',
                                    'Kelurahan', 'Address','Branch', 'Cluster SCS','Antenna Type', 'Antenna Height', 'Azimut', 
                                    'M-Tilt Antena','E-Tilt Antena', 'Tower Provider',"LAC", "RNC Name", "RAC", "URA ID", 
                                    "SAC", "SC","MCC", "MNC", "F1/F2/F3/U900"]
                
                df_raw13.insert(0, 'network_type_band', "3G", allow_duplicates=False)
                df_raw13.insert(0, 'vendor', "ZTE", allow_duplicates=False)
                df_raw13.insert(0, 'order_id', np.nan, allow_duplicates=False)
                
                df_raw13.insert(6, 'Site Name OSS', np.nan, allow_duplicates=False)
                
                df_raw13.insert(9, 'SoW', np.nan, allow_duplicates=False)
        
                df_raw13.insert(17, 'Network Area', "Area 04", allow_duplicates=False)
                df_raw13.insert(17, 'Regional', "R11 Papua Maluku", allow_duplicates=False)
                df_raw13.insert(17, 'LOC No', np.nan, allow_duplicates=False)
                
                df_raw13.insert(22, 'Config', np.nan, allow_duplicates=False)
                df_raw13.insert(22, 'Type BTS', np.nan, allow_duplicates=False)
                df_raw13.insert(22, 'NE Type', np.nan, allow_duplicates=False)
                df_raw13.insert(22, 'Time Zone', np.nan, allow_duplicates=False)
                
                df_raw13.insert(34, 'SPC', np.nan, allow_duplicates=False)
                
                df_raw13.insert(39, 'ID', "-", allow_duplicates=False)
                df_raw13.insert(39, 'MGW', np.nan, allow_duplicates=False)
                df_raw13.insert(39, 'MSS SPC', "Pooling", allow_duplicates=False)
                df_raw13.insert(39, 'MSS', "Pooling", allow_duplicates=False)
                df_raw13.insert(39, 'MSS Name', "Pooling", allow_duplicates=False)
                df_raw13.insert(39, 'WBTS ID	', "TBD", allow_duplicates=False)
        
                df_raw13.insert(47, 'Frequency', np.nan, allow_duplicates=False)		
                
                df_raw13.loc[:,'Remark'] = ty
                
                df_raw3 = pd.concat([df_raw3,df_raw13],axis = 0,ignore_index= True)
                
        df_raw3['Site Name OSS'] = df_raw3['BTS/NodeB/eNodeB_Name'].str[2:]
        df_raw3['NE Type'] = df_raw3['ne_id'].str[6]
    
        df_3GTPASS = pd.DataFrame()
        
        for x in range(9):
            ty = x + 1      
            df_selection3 = pd.DataFrame()
            df_selection3 = df_raw3[df_raw3['Remark'] == ty].copy()
            df_colon13 = df_selection3[['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena', "SC"]].copy()
            df_colon3 = pd.DataFrame()
            df_3Gprep = pd.DataFrame()
            if not df_selection3.empty:
                if ty == 1:
                    df_3GTPASS1 = df_selection3.copy()
                    df_3GTPASS = pd.concat([df_3GTPASS,df_3GTPASS1],axis = 0,ignore_index= True)
                else:
                    for z in range(0,len(df_selection3),ty):
                        df_tempc = pd.DataFrame()
                        for y in range(ty):
                            t = z + y
                            data = df_colon13.loc[df_colon13.index[t],['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena', "SC"]].values
                            df_tempc.insert(y, y, data, allow_duplicates=False)
                        if ty == 1:
                            df_tempc['Concat'] = df_tempc[0].astype(str)
                        else:
                            df_tempc['Sementara Concat'] = df_tempc[0].astype(str)
                            for w in range(1,ty):
                                df_tempc['Concat'] = df_tempc['Sementara Concat'] +';'+ df_tempc[w].astype(str)
                                df_tempc['Sementara Concat'] = df_tempc['Concat']
                        df_tempt = df_tempc.transpose()
                        df_tempt.columns =['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena', "SC"]
                        df_tempt.loc['Concat','ne_id'] = df_tempt.loc[0,'ne_id']
                        df_tempc =  df_tempt.loc['Concat',:].copy()
                        df_colon3 = df_colon3.append(df_tempc,ignore_index=True)
                        df_3Gprept = pd.DataFrame(df_selection3.loc[df_selection3.index[(z)],df_selection3.columns].values.reshape(1,-1), columns = df_selection3.columns)
                        df_3Gprep = df_3Gprep.append(df_3Gprept,ignore_index=True)
                        df_colon3 = df_colon3.drop(columns=['ne_id'],axis = 1)
                df_3Gprep.update(df_colon3)
                df_3GTPASS = pd.concat([df_3GTPASS,df_3Gprep],axis = 0,ignore_index= True)
    
        df_3GTPASS['NE Type'] = df_3GTPASS['NE Type'].replace({"M":"Macro","I":"Indoor","H":"Hotel"})
        df_3GTPASS['Config'] = df_3GTPASS['Remark'].copy()
        df_3GTPASS['Config'] = df_3GTPASS['Config'].replace({1:'S1', 2:'S11', 3:'S111', 4:'S1111', 5:'S11111', 6:'S111111'})
        df_3GTPASS = df_3GTPASS.drop('Remark',axis =1)
        df_3GTPASS = pd.merge(df_3GTPASS,df_rncspc, how = 'left' , on = 'RNC Name')
        df_3GTPASS["SPC"] = df_3GTPASS["RNC SPC"].copy()
        df_3GTPASS = df_3GTPASS.drop("RNC SPC",axis = 1)
        df_3GTPASS = pd.merge(df_3GTPASS,df_mgw, how = 'left' , on = 'RNC Name')
        df_3GTPASS["MGW_x"] = df_3GTPASS["MGW_y"]
        df_3GTPASS = df_3GTPASS.rename(columns = {'MGW_x': 'MGW'}, inplace = False)
        df_3GTPASS = df_3GTPASS.drop("MGW_y",axis = 1)
        df_3GTPASS["ID"] = df_3GTPASS["Cell ID"]
        df_3GTPASS["Frequency"] = df_3GTPASS['ne_id'].str[7]
        df_3GTPASS['Frequency'] = df_3GTPASS['Frequency'].replace({'X': 2100,'W': 2100,'Z': 2100,'A': 900 })
        
        df_3GTPASS['Kabupaten'] = df_3GTPASS['Kabupaten'].str.title()
        df_3GTPASS = pd.merge(df_3GTPASS,df_tz, how = 'left' , on = 'Kabupaten')
        df_3GTPASS["Time Zone_x"] = df_3GTPASS["Time Zone_y"]
        df_3GTPASS = df_3GTPASS.rename(columns = {'Time Zone_x': 'Time Zone'}, inplace = False)
        df_3GTPASS = df_3GTPASS.drop("Time Zone_y",axis = 1)
        
        df_3GTPASS = pd.merge(df_3GTPASS,df_locno1, how = 'left' , on = 'Kabupaten')
        df_3GTPASS["LOC No_x"] = df_3GTPASS["LOC No_y"]
        df_3GTPASS = df_3GTPASS.rename(columns = {'LOC No_x': 'LOC No'}, inplace = False)
        df_3GTPASS = df_3GTPASS.drop("LOC No_y",axis = 1)
        
        
        sheetname1 = '3G'
        df_3GTPASS.to_excel(writer, sheet_name = sheetname1, index=False)
    
    
    
    if "2G" in wb1.sheetnames:
        df_nodin2G = pd.read_excel(filename, sheet_name="2G",header = 2, engine= 'openpyxl',)
        
        df_2Gcount= pd.DataFrame()
        df_2Gcount['NE ID'] = df_nodin2G['NE ID'].copy()
        df_2Gcount = df_2Gcount.groupby(['NE ID'])[['NE ID']].count()
        
        df_raw2 = pd.DataFrame()
    
        for x in range(9):
            ty = x + 1       
            df_selection2 = pd.DataFrame()
            df_selection2['NE ID'] = df_2Gcount[df_2Gcount['NE ID'] == ty].index
            
            if not df_selection2.empty:
                df_temp2G = pd.merge(df_selection2,df_nodin2G, how = 'left' , on = 'NE ID')
                df_raw12 = df_temp2G[["SITE ID", "NE ID", "SITE NAME\n(CONNECTED)", "CI Plan Sector", "CELL NAME\n(CONNECTED)", "Long", "Lat", "PROVINCE", "KABUPATEN", 
                                    "KECAMATAN", "KELURAHAN", "Address" ,"LOCNO", "BRANCH",  "Sales cluster", "ANTENNA TYPE", "Antenna Height\n(Implementation)", 
                                    "AFTER.1", "MDT", "EDT", "TP", "LAC","BSC", "MGW SPC", "RAC", "MGW SPC","MGW"
                                    ]].copy()
                df_raw12.columns = ['site_id', 'ne_id', 'BTS/NodeB/eNodeB_Name','Cell ID','Cell Name', 'Longitude', 'Latitude', 'Province', 'Kabupaten', 'Kecamatan', 
                                    'Kelurahan', 'Address', 'LOC No', 'Branch', 'Cluster SCS','Antenna Type', 'Antenna Height', 'Azimut', 
                                    'M-Tilt Antena', 'E-Tilt Antena', 'Tower Provider', "LAC", "BSC/RNC Name", "BSC SPC", "RAC","MSS SPC", "MSS Name"]
                
                df_raw12.insert(0, 'network_type_band', "2G", allow_duplicates=False)
                df_raw12.insert(0, 'vendor', "ZTE", allow_duplicates=False)
                df_raw12.insert(0, 'order_id', np.nan, allow_duplicates=False)
                
                df_raw12.insert(6, 'Site Name OSS', np.nan, allow_duplicates=False)
                
                df_raw12.insert(9, 'SoW', np.nan, allow_duplicates=False)
        
                df_raw12.insert(18, 'Network Area', "Area 04", allow_duplicates=False)
                df_raw12.insert(18, 'Regional', "R11 Papua Maluku", allow_duplicates=False)
                
                df_raw12.insert(22, 'Config', np.nan, allow_duplicates=False)
                df_raw12.insert(22, 'Type BTS', np.nan, allow_duplicates=False)
                df_raw12.insert(22, 'NE Type', np.nan, allow_duplicates=False)
                df_raw12.insert(22, 'Time Zone', np.nan, allow_duplicates=False)
                
                df_raw12.insert(36, 'URA ID', "TBD", allow_duplicates=False)
                df_raw12.insert(38, 'WBTS UD', "TBD", allow_duplicates=False)
                df_raw12.insert(40, 'MSS', "MSS POLL", allow_duplicates=False)
                
                df_raw12.loc[:,'Remark'] = ty
                
                df_raw2 = pd.concat([df_raw2,df_raw12],axis = 0,ignore_index= True)
                
        
        df_raw2['Site Name OSS'] = df_raw2['BTS/NodeB/eNodeB_Name'].str[2:]
        df_raw2['NE Type'] = df_raw2['ne_id'].str[6]
        
        
        
        df_2GTPASS = pd.DataFrame()
        
        for x in range(9):
            ty = x + 1   
            df_selection2 = pd.DataFrame()
            df_selection2 = df_raw2[df_raw2['Remark'] == ty].copy()
            df_colon12 = df_selection2[['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena']].copy()
            df_colon2 = pd.DataFrame()
            df_2Gprep = pd.DataFrame()
            if not df_selection2.empty:
                if ty == 1:
                    df_2GTPASS1 = df_selection2.copy()
                    df_2GTPASS = pd.concat([df_2GTPASS,df_2GTPASS1],axis = 0,ignore_index= True)
                else:
                    for z in range(0,len(df_selection2),ty):
                        df_tempc = pd.DataFrame()
                        for y in range(ty):
                            t = z + y
                            data = df_colon12.loc[df_colon12.index[t],['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena']].values
                            df_tempc.insert(y, y, data, allow_duplicates=False)
                        if ty == 1:
                            df_tempc['Concat'] = df_tempc[0].astype(str)
                        else:
                            df_tempc['Sementara Concat'] = df_tempc[0].astype(str)
                            for w in range(1,ty):
                                df_tempc['Concat'] = df_tempc['Sementara Concat'] +';'+ df_tempc[w].astype(str)
                                df_tempc['Sementara Concat'] = df_tempc['Concat']
                        df_tempt = df_tempc.transpose()
                        df_tempt.columns =['ne_id','Cell ID','Cell Name','Azimut','M-Tilt Antena', 'E-Tilt Antena']
                        df_tempt.loc['Concat','ne_id'] = df_tempt.loc[0,'ne_id']
                        df_tempc =  df_tempt.loc['Concat',:].copy()
                        df_colon2 = df_colon2.append(df_tempc,ignore_index=True)
                        df_2Gprept = pd.DataFrame(df_selection2.loc[df_selection2.index[(z)],df_selection2.columns].values.reshape(1,-1), columns = df_selection2.columns)
                        df_2Gprep = df_2Gprep.append(df_2Gprept,ignore_index=True)
                        df_colon2 = df_colon2.drop(columns=['ne_id'],axis = 1)
                df_2Gprep.update(df_colon2)
                df_2GTPASS = pd.concat([df_2GTPASS,df_2Gprep],axis = 0,ignore_index= True)
        
        df_2GTPASS['NE Type'] = df_2GTPASS['NE Type'].replace({"M":"Macro","I":"Indoor","H":"Hotel"})
        df_2GTPASS['Config'] = df_2GTPASS['Remark'].copy()
        df_2GTPASS['Config'] = df_2GTPASS['Config'].replace({1:'S1', 2:'S11', 3:'S111', 4:'S1111', 5:'S11111', 6:'S111111'})
        df_2GTPASS = df_2GTPASS.drop('Remark',axis =1)
        
        df_2GTPASS['Kabupaten'] = df_2GTPASS['Kabupaten'].str.title()
        df_2GTPASS = pd.merge(df_2GTPASS,df_tz, how = 'left' , on = 'Kabupaten')
        df_2GTPASS["Time Zone_x"] = df_2GTPASS["Time Zone_y"]
        df_2GTPASS = df_2GTPASS.rename(columns = {'Time Zone_x': 'Time Zone'}, inplace = False)
        df_2GTPASS = df_2GTPASS.drop("Time Zone_y",axis = 1)
        
        #save to spesific directory
        
        sheetname1 = '2G'
        df_2GTPASS.to_excel(writer, sheet_name = sheetname1, index=False)
    
    
    writer.save()
    fileo_path = os.path.join(directory, file_output)
    return([file_output,fileo_path])
